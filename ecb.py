# 📊 수입지출 자동화 보고서 (fillna 안정 처리 포함 최종)
import pandas as pd
import io, re
from google.colab import files
from IPython.display import display, HTML, clear_output
import ipywidgets as widgets

mapping_dict = {
    "지정후원금(지역)": ("지정후원금(지역)", 33),
    "지정후원금(기타)": ("지정후원금(기타)", 34),
    "전년도이월금(지정후원금기타)": ("전년도이월금(지정후원금기타)", 82),
    "전년도이월금(법인후원금_지역기타)": ("전년도이월금(법인후원금_지역기타)", 66),
    "전년도이월금(법인후원금_지역)": ("전년도이월금(법인후원금_지역)", 65),
    "전년도이월금(후원금)": ("전년도 이월금(후원금)", 81),
    "전년도이월금(보조금)": ("전년도 이월금(보조금)", 71),
    "전년도이월금(법인후원금기타)": ("전년도 이월금(법인후원금기타)", 62),
    "전년도이월금(법인후원금)": ("전년도 이월금(법인)", 61),
    "사업수입": ("사업수입", 42),
    "법인전입금(후원금)지역(기타)": ("법인전입금(후원금)지역(기타)", 26),
    "법인전입금(후원금)지역": ("법인전입금(후원금)지역", 25),
    "법인전입금(후원금)기타": ("법인전입금(후원금) 기타", 22),
    "기타예금이자(후원금)": ("기타예금이자(후원금)", 46),
    "기타예금이자": ("기타예금이자", 45),
    "과년도수입": ("없음", 99),
    "후원물품(지부)": ("수증물품", 99),
    "후원물품(본부)": ("수증물품", 99),
    "지정후원금(본부)": ("지정후원금", 32),
    "지정후원금": ("지정후원금", 32),
    "전년도이월액(보육료)": ("전년도 이월금(기타)", 91),
    "전년도이월금(사업수입및기타)": ("전년도 이월금(기타)", 91),
    "전년도이월금(사업수입)": ("전년도 이월금(기타)", 91),
    "전년도이월금(기타)": ("전년도 이월금(기타)", 91),
    "이자수입(후원금)": ("이자수입(본부)", 47),
    "이자수입(퇴직연금운용수익)": ("이자수입(본부)", 47),
    "이자수입(본부)": ("이자수입(본부)", 47),
    "인건비(시도보조금)": ("시도보조금", 13),
    "운영비(시도보조금)": ("시도보조금", 13),
    "시도보조금": ("시도보조금", 13),
    "사업비(시도보조금)": ("시도보조금", 13),
    "인건비(시군구보조금)": ("시군구보조금", 14),
    "운영비(시군구보조금)": ("시군구보조금", 14),
    "시군구보조금": ("시군구보조금", 14),
    "사업비(시군구보조금)": ("시군구보조금", 14),
    "특별활동비(어린이집)": ("사용자부담금", 43),
    "연장보육료": ("사용자부담금", 43),
    "사용자부담금(본부결산용)": ("사용자부담금", 43),
    "사용자부담금": ("사용자부담금", 43),
    "부모부담 보육료(어린이집)": ("사용자부담금", 43),
    "기타필요경비(어린이집)": ("사용자부담금", 43),
    "시설후원금": ("비지정후원금", 31),
    "비지정후원금": ("비지정후원금", 31),
    "직장어린이집위탁전입금": ("보조금", 16),
    "정부지원 보육료(어린이집)": ("보조금", 16),
    "자본보조금": ("보조금", 16),
    "인건비 보조금": ("보조금", 16),
    "보조금": ("보조금", 16),
    "정기후원금": ("법인전입금(후원금)", 21),
    "일시후원금(본부)": ("법인전입금(후원금)", 21),
    "법인전입금(후원금)": ("법인전입금(후원금)", 21),
    "적립금처분수입(어린이집)": ("잡수입", 41),
    "잡수익": ("잡수입", 41),
    "기타잡수입": ("잡수입", 41),
    "기타지원금": ("기타보조금", 15),
    "기타 보조금": ("기타보조금", 15),
    "인건비(국고보조금)": ("국고보조금", 12),
    "운영비(국고보조금)": ("국고보조금", 12),
    "사업비(국고보조금)": ("국고보조금", 12),
    "국고보조금": ("국고보조금", 12),
    "잡수입": ("잡수입", 41),
    "기타보조금": ("기타보조금", 15),
    "전년도이월금(법인)": ("전년도 이월금(법인)", 61),
    "": ("매핑없음", 999),
    "수증물품": ("수증물품", 99),
    "이이월금": ("이이월금", 999),
    "합계": ("합계", 999),
}

def flatten_multi_header(file_io):
    headers = pd.read_excel(file_io, nrows=2, header=None)
    if hasattr(file_io, 'seek'): file_io.seek(0)
    for i in range(headers.shape[1]):
        for r in range(headers.shape[0]):
            if re.match(r"Unnamed: \\d+", str(headers.iloc[r, i])):
                headers.iloc[r, i] = None
    row0, merged0, current = list(headers.iloc[0]), [], ""
    for val in row0:
        if pd.notna(val): current = val
        merged0.append(current)
    row1 = list(headers.iloc[1])
    merged1 = [val if pd.notna(val) else merged0[i] for i, val in enumerate(row1)]
    new_columns = [f"{top}_{bottom}" if bottom and top != bottom else top for top, bottom in zip(merged0, merged1)]
    df_data = pd.read_excel(file_io, skiprows=2, header=None)
    df_data.columns = new_columns
    return df_data

def normalize_key(x):
    return str(x).replace(" ", "").replace("\u200b", "").strip()


def map_semo_column(df):
    semo_col = next((col for col in df.columns if "세목" in col), None)
    if not semo_col:
        raise ValueError("❌ '세목' 컬럼을 찾을 수 없습니다.")
    idx = df.columns.get_loc(semo_col) + 1
    df.insert(idx, "재원", df[semo_col].map(lambda x: mapping_dict.get(normalize_key(x), ("매핑없음", 999))[0]))
    df.insert(idx + 1, "재원순서", df[semo_col].map(lambda x: mapping_dict.get(normalize_key(x), ("매핑없음", 999))[1]))
    return df



def insert_subtotals(df, group_col, value_cols):
    subtotal_rows = []
    for key, group in df.groupby(group_col, sort=False):
        subtotal = group[value_cols].sum(numeric_only=True)
        subtotal_row = {col: None for col in df.columns}
        subtotal_row[group_col] = f"{key} 소계"
        for col in value_cols:
            subtotal_row[col] = subtotal[col]
        subtotal_rows.append((group.index[-1] + 0.1, subtotal_row))
    for idx, row in sorted(subtotal_rows, reverse=True):
        df.loc[idx] = row
    df = df.sort_index().reset_index(drop=True)
    return df



def compose_jaewon_detail(df):
    def normalize_key(x):
        return str(x).replace(" ", "").replace("\u200b", "").strip()

    jaewon_col = next((col for col in df.columns if "예산정보_재원" in col), None)
    hang_col = next((col for col in df.columns if "예산정보_항" in col), None)
    if not jaewon_col or not hang_col:
        raise ValueError("❌ '예산정보_항' 또는 '예산정보_재원' 컬럼을 찾을 수 없습니다.")

    df = df.rename(columns={jaewon_col: "재원", hang_col: "항"})
    항_idx = df.columns.get_loc("항") + 1
    df.insert(항_idx, "재원", df.pop("재원"))
    df.insert(
        항_idx + 1,
        "재원순서",
        df["재원"].map(lambda x: mapping_dict.get(normalize_key(x), ("매핑없음", 999))[1])
    )
    df.insert(
        항_idx + 2,
        "재원항",
        df["항"].astype(str).str.strip() + "_" + df["재원"].astype(str).str.strip()
    )
    return df

def clean_numeric_columns(df, start_col):
    idx = df.columns.get_loc(start_col) + 1
    num_cols = df.columns[idx:]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "").str.strip(), errors="coerce")
    return df, list(num_cols)

upload_expense = widgets.FileUpload(description="📤 지출 업로드", accept='.xls,.xlsx', multiple=False)
upload_income = widgets.FileUpload(description="📥 수입 업로드", accept='.xls,.xlsx', multiple=False)
run_button = widgets.Button(description="🚀 보고서 생성", button_style="success")
output = widgets.Output()

display(HTML("<h4>1️⃣ 지출 파일 업로드</h4>")); display(upload_expense)
display(HTML("<h4>2️⃣ 수입 파일 업로드</h4>")); display(upload_income)
display(run_button); display(output)




# ✅ 시트8: 시트7 요약 데이터 기준 누적합 시트 생성 함수 정의
def create_cumulative_sheet(df_diff):
    df = df_diff.copy()
    df = df.sort_values(["기관", "재원", "재원순서"])

    cum_df = df[["기관", "재원", "재원순서"]].copy()

    for mode in ["예산(계획)", "집행"]:
        for month in range(1, 13):
            cols = [f"{m}월_{mode}" for m in range(1, month + 1)]
            valid_cols = [c for c in cols if c in df.columns]
            cum_col = f"{month}월_누적_{mode}"
            cum_df[cum_col] = df[valid_cols].sum(axis=1)

    return cum_df



def run_final_report(b):
    with output:
        clear_output()
        print("🚦 보고서 생성 시작...")

        if not upload_expense.value:
            print("❌ 지출 파일을 먼저 업로드하세요."); return
        if not upload_income.value:
            print("❌ 수입 파일을 먼저 업로드하세요."); return

        print("📤 지출 파일 정제 중...")
        df_expense_raw = flatten_multi_header(io.BytesIO(list(upload_expense.value.values())[0]['content']))
        df_expense = compose_jaewon_detail(df_expense_raw)
        df_expense, expense_cols = clean_numeric_columns(df_expense, "예산정보_세목")
        print("✅ 지출 정제 완료")

        print("📥 수입 파일 정제 중...")
        df_income_raw = flatten_multi_header(io.BytesIO(list(upload_income.value.values())[0]['content']))
        df_income = map_semo_column(df_income_raw)
        df_income, income_cols = clean_numeric_columns(df_income, "재원순서")
        print("✅ 수입 정제 완료")

        print("📊 피벗 및 병합 처리 중...")
        summary_expense_1 = df_expense.groupby(["기관", "항", "재원", "재원항", "재원순서"])[expense_cols].sum().reset_index()

        print("📊 피벗 및 병합 처리 중...")

        # ✅ 시트3용 항-재원 피벗
        summary_expense_1 = df_expense.groupby(["기관", "항", "재원", "재원항", "재원순서"])[expense_cols].sum().reset_index()

        # ✅ 시트4용 지출 재원피벗 생성
        summary_expense_2 = df_expense.groupby(["기관", "재원", "재원순서"])[expense_cols].sum().reset_index()
        summary_expense_2.insert(1, "수입지출", "지출")
        summary_expense_2 = summary_expense_2.sort_values(["기관", "수입지출", "재원", "재원순서"])

        # ✅ 시트5: 수입 피벗 요약
        summary_income_filtered = df_income[df_income["재원"] != "매핑없음"].copy()

        summary_income_5 = (
            summary_income_filtered
            .groupby(["기관", "재원", "재원순서"])[income_cols]
            .sum()
            .reset_index()
        )
        summary_income_5.insert(1, "수입지출", "수입")
        summary_income_5 = summary_income_5.sort_values(["기관", "재원", "재원순서"])



        # ✅ 시트6: 수직 병합
        common_cols_no_항 = ["기관", "수입지출", "재원", "재원순서"]

        # 기준이 되는 정렬된 컬럼 (시트5 기준 우선)
        ordered_cols = [c for c in income_cols if c in df_income.columns]
        ordered_cols += [c for c in expense_cols if c in df_expense.columns and c not in ordered_cols]

        summary_merged = pd.concat([
            summary_income_5[common_cols_no_항 + [c for c in ordered_cols if c in summary_income_5.columns]],
            summary_expense_2[common_cols_no_항 + [c for c in ordered_cols if c in summary_expense_2.columns]]
        ], axis=0, ignore_index=True)

        summary_merged = summary_merged.sort_values(["기관", "수입지출", "재원", "재원순서"])




        # ✅ 시트7: 수입지출 차액 요약 (5번 - 4번 기준)
        target_cols = []
        for month in range(1, 13):
            m = f"{month}월"
            target_cols.append(f"{m}_예산(계획)")
            target_cols.append(f"{m}_집행")

        income_trim = summary_income_5[["기관", "재원", "재원순서"] + [col for col in target_cols if col in summary_income_5.columns]].copy()
        expense_trim = summary_expense_2[["기관", "재원", "재원순서"] + [col for col in target_cols if col in summary_expense_2.columns]].copy()

        summary_diff_only = pd.merge(
            income_trim,
            expense_trim,
            on=["기관", "재원", "재원순서"],
            suffixes=("_수입", "_지출"),
            how="outer"
        )

        for col in target_cols:
            col_income = f"{col}_수입"
            col_expense = f"{col}_지출"
            summary_diff_only[col] = (
                summary_diff_only.get(col_income, pd.Series([0] * len(summary_diff_only))).fillna(0)
                - summary_diff_only.get(col_expense, pd.Series([0] * len(summary_diff_only))).fillna(0)
            )

        summary_diff_only = summary_diff_only[["기관", "재원", "재원순서"] + target_cols]
        summary_diff_only = summary_diff_only.sort_values(["기관", "재원", "재원순서"])

        # ✅ 시트8: 시트7 요약 데이터 기준 누적합 시트 생성
        summary_cumulative = create_cumulative_sheet(summary_diff_only)


        # ✅ 시트9_관-세목 / 지출 피벗 생성
        pivot_expense_9 = df_expense.groupby(
            ["기관", "예산정보_관", "항", "예산정보_목", "예산정보_세목"]
        )[expense_cols].sum().reset_index()

        # ✅ 시트10_관-세목-재원 / 지출 피벗 생성
        pivot_expense_10 = df_expense.groupby(
            ["기관", "예산정보_관", "항", "예산정보_목", "예산정보_세목", "재원", "재원순서"]
        )[expense_cols].sum().reset_index()

        # ✅ 시트11_수입_관-세목-재원 피벗 생성 + 정렬(기관, 재원순서)
        pivot_income_11 = df_income.groupby(
            ["기관", "예산정보_관", "예산정보_항", "예산정보_목", "예산정보_세목", "재원", "재원순서"]
        )[income_cols].sum().reset_index()
        pivot_income_11 = pivot_income_11.sort_values(["기관", "재원순서"])

        # ✅ 컬럼명 통일: '사업장_사업유형' → '사업유형'
        if "사업장_사업유형" in df_income.columns:
            df_income.rename(columns={"사업장_사업유형": "사업유형"}, inplace=True)

        if "사업장_사업유형" in df_expense.columns:
            df_expense.rename(columns={"사업장_사업유형": "사업유형"}, inplace=True)

        # ✅ 시트12: 수입 요약 - 사업유형, 사업장명, 재원 기준 그룹바이
        summary_income_by_fund = df_income.groupby(
            ["사업유형", "사업장명", "재원"]
        )[[col for col in income_cols if col in df_income.columns]].sum().reset_index()

        # ✅ 재원 우선순위 정렬
        fund_priority = [
            "경상보조금", "기타보조금", "후원금", "후원물품",
            "법인전입금(후원금)기타", "법인전입금(지역)", "법인전입금(후원금)", "잡수입", "이월금"
        ]
        summary_income_by_fund["정렬순서"] = summary_income_by_fund["재원"].apply(
            lambda x: fund_priority.index(x) if x in fund_priority else 999
        )
        summary_income_by_fund = summary_income_by_fund.sort_values(
            ["사업유형", "사업장명", "정렬순서", "재원"]
        ).drop(columns="정렬순서")

        # ✅ 시트13: 지출 요약 - 사업유형, 사업장명, 항 기준 그룹바이
        summary_expense_by_hang = df_expense.groupby(
            ["사업유형", "사업장명", "항"]
        )[[col for col in expense_cols if col in df_expense.columns]].sum().reset_index()

        # ✅ 항목 우선순위 정렬
        hang_priority = [
            "인건비", "업무추진비", "운영비", "시설비", "사업비", "잡지출", "기타"
        ]
        summary_expense_by_hang["정렬순서"] = summary_expense_by_hang["항"].apply(
            lambda x: hang_priority.index(x) if x in hang_priority else 999
        )
        summary_expense_by_hang = summary_expense_by_hang.sort_values(
            ["사업유형", "사업장명", "정렬순서", "항"]
        ).drop(columns="정렬순서")

        # ✅ 시트12: 사업유형별 수입 소계 삽입
        income_value_cols = [
            col for col in summary_income_by_fund.columns
            if col not in ["사업유형", "사업장명", "재원"]
        ]
        summary_income_by_fund = insert_subtotals(
            summary_income_by_fund, "사업유형", income_value_cols
        )

        # ✅ 시트13: 사업유형별 지출 소계 삽입
        expense_value_cols = [
            col for col in summary_expense_by_hang.columns
            if col not in ["사업유형", "사업장명", "항"]
        ]
        summary_expense_by_hang = insert_subtotals(
            summary_expense_by_hang, "사업유형", expense_value_cols
        )

       

        # ✅ 불필요한 컬럼 제거
        # ✅ 시트1_지출재원항매핑: 수입지출계획_증감사유만 제거, 나머지는 유지
        if "수입지출계획_증감사유" in df_expense.columns:
            df_expense.drop(columns=["수입지출계획_증감사유"], inplace=True)


        # ✅ 시트2_수입재원매핑: 불필요한 컬럼 제거
        df_income.drop(
            columns=[c for c in ["수입지출계획_증감사유", "예산정보_재원"] if c in df_income.columns],
            inplace=True
        )

        # ✅ 시트3/4: 지출 피벗 요약본들 정리
        for df in [summary_expense_1, summary_expense_2]:
            df.drop(
                columns=[c for c in ["수입지출계획_증감사유", "예산정보_목", "예산정보_세목"] if c in df.columns],
                inplace=True
            )

        # ✅ 시트5_수입재원피벗: 불필요한 컬럼 제거
        summary_income_5.drop(
            columns=[c for c in ["수입지출계획_증감사유", "예산정보_재원"] if c in summary_income_5.columns],
            inplace=True
        )

        # ✅ 시트6_통합수입지출: 불필요한 컬럼 제거
        summary_merged.drop(
            columns=[c for c in ["수입지출계획_증감사유", "예산정보_목", "예산정보_세목", "예산정보_재원"] if c in summary_merged.columns],
            inplace=True
        )


        # ✅ 저장
        save_path = "/content/수지_경리보고서월별(통합).xlsx"

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            df_expense.to_excel(writer, index=False, sheet_name="시트1_지출재원항매핑")
            df_income.to_excel(writer, index=False, sheet_name="시트2_수입재원매핑")
            summary_expense_1.to_excel(writer, index=False, sheet_name="시트3_지출 항-재원 피벗")
            summary_expense_2.to_excel(writer, index=False, sheet_name="시트4_지출재원피벗")
            summary_income_5.to_excel(writer, index=False, sheet_name="시트5_수입재원피벗")
            summary_merged.to_excel(writer, index=False, sheet_name="시트6_통합수입지출")
            summary_diff_only.to_excel(writer, index=False, sheet_name="시트7_차액요약")
            summary_cumulative.to_excel(writer, index=False, sheet_name="시트8_누적합요약")
            # ✅ 추가: 시트9~11 저장
            pivot_expense_9.to_excel(writer, index=False, sheet_name="시트9_지출_관-세목")
            pivot_expense_10.to_excel(writer, index=False, sheet_name="시트10_지출_관-세목-재원")
            pivot_income_11.to_excel(writer, index=False, sheet_name="시트11_수입_관-세목-재원")
            summary_income_by_fund.to_excel(writer, index=False, sheet_name="시트12_수입요약_재원기준")
            summary_expense_by_hang.to_excel(writer, index=False, sheet_name="시트13_지출요약_항목기준")




        # ✅ 정렬
        sort_targets = {
            "시트1_지출재원항매핑": ["기관", "예산정보_사업구분", "예산정보_세부사업", "예산정보_관", "항", "재원순서"],
            "시트2_수입재원매핑": ["기관", "예산정보_항", "재원순서"],
            "시트3_지출 항-재원 피벗": ["기관", "항", "재원순서"],
            "시트4_지출재원피벗": ["기관", "재원순서"],
            "시트5_수입재원피벗": ["기관", "재원순서"],
            "시트6_통합수입지출": ["기관", "수입지출", "재원순서"],
            "시트7_차액요약": ["기관", "재원순서"],
            "시트8_누적합요약": ["기관", "재원순서"],
            # ✅ 추가: 시트9~11
            "시트9_지출_관-세목": ["기관", "예산정보_관", "항", "예산정보_목", "예산정보_세목"],
            "시트10_지출_관-세목-재원": ["기관", "예산정보_관", "항", "예산정보_목", "예산정보_세목", "재원순서"],
            "시트11_수입_관-세목-재원": ["기관", "재원순서"],
        }



        # ✅ 정렬 후 숫자만 콤마 서식 적용 (2행부터 전부 검사)
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(save_path)

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center")

        for i in range(1, 12):
            sheet_prefix = f"시트{i}_"
            matched_sheets = [s for s in wb.sheetnames if s.startswith(sheet_prefix)]
            if not matched_sheets:
                continue
            ws = wb[matched_sheets[0]]

            col_widths = {}

            # 🔹 헤더 스타일 + 열 너비 계산 포함
            for idx, cell in enumerate(ws[1], start=1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                val = str(cell.value) if cell.value is not None else ""
                width = len(val.encode("utf-8")) // 2 + 2
                col_letter = get_column_letter(idx)
                col_widths[col_letter] = width

            # 🔹 숫자만 콤마 서식 + 데이터 너비 계산 (2행부터 전부 검사)
            for row in ws.iter_rows(min_row=2):
                for idx, cell in enumerate(row, start=1):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
                    val = str(cell.value) if cell.value is not None else ""
                    width = len(val.encode("utf-8")) // 2 + 2
                    col_letter = get_column_letter(idx)
                    col_widths[col_letter] = max(col_widths.get(col_letter, 0), width)

            # 🔹 열 너비 반영
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

        wb.save(save_path)





        # ✅ 조건부 서식 (시트7)
        import openpyxl
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import CellIsRule

        wb = openpyxl.load_workbook(save_path)
        ws = wb["시트7_차액요약"]

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        min_row, max_row = 2, ws.max_row
        min_col, max_col = 4, ws.max_column

        for col in range(min_col, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            col_range = f"{col_letter}{min_row}:{col_letter}{max_row}"
            ws.conditional_formatting.add(col_range, CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))
            ws.conditional_formatting.add(col_range, CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill))

        wb.save(save_path)


        # ✅ 조건부 서식 (시트8)
        ws8 = wb["시트8_누적합요약"]

        min_row, max_row = 2, ws8.max_row
        min_col, max_col = 4, ws8.max_column

        for col in range(min_col, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            col_range = f"{col_letter}{min_row}:{col_letter}{max_row}"
            ws8.conditional_formatting.add(col_range, CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))
            ws8.conditional_formatting.add(col_range, CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill))

        wb.save(save_path)



        print("✅ 완료! 보고서를 다운로드하세요 👇")
        files.download(save_path)





run_button.on_click(run_final_report)
