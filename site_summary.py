# site_summary.py

import os
import pandas as pd
import ipywidgets as widgets
from IPython.display import display, clear_output
from google.colab import files
import openpyxl
from openpyxl.styles import PatternFill

# 공용 출력 영역
output = widgets.Output()

# UI 구성요소
upload_button = widgets.Button(description="📤 파일 업로드", button_style="info")
explore_button = widgets.Button(description="🔍 파일 탐색", button_style="warning")
filename_dropdown = widgets.Dropdown(description="📄 파일 선택:", layout=widgets.Layout(width="60%"))
run_button = widgets.Button(description="✅ 진행하기", button_style="success", disabled=True)


# -------- 공통 유틸 --------
def get_priority(value: str) -> int:
    if value == "총계":
        return 99
    elif value == "소계":
        return 9
    return 1

def sort_postprocessed_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["sort_priority"] = df["기관"].apply(get_priority)
    df = df.sort_values(
        by=["사업장_사업유형", "sort_priority", "사업장_권역", "기관"]
    ).drop(columns=["sort_priority"]).reset_index(drop=True)
    return df


# -------- 데이터 처리 함수 --------
def summarize_expense_by_account(df_expense: pd.DataFrame) -> pd.DataFrame:
    hang_mapping = {
        "사업비": "사업비", "모금관리비": "기타", "운영비": "운영비", "조사연구사업": "기타",
        "인건비": "인건비", "전출금": "기타", "예비비및기타": "기타", "시설비": "시설비",
        "업무추진비": "업무추진비", "반환금": "기타", "교육비": "사업비", "운영비(시설)": "시설비",
        "부채상환금": "기타", "기본보육활동": "사업비", "원장인건비": "인건비",
        "보육교직원인건비": "인건비", "기관부담금": "기타", "수익자부담경비": "기타"
    }
    df = df_expense.copy()
    df["기관"] = df.apply(lambda row: row["사업장_예산부서"] if row["사업장_권역"] == "본부" else row["기관"], axis=1)
    df["매핑"] = df["항"].map(hang_mapping)

    summary = df.groupby(["사업장_사업유형", "사업장_권역", "기관"], as_index=False)[
        ["예산(계획)_전기(C)", "예산(계획)_당기(B)"]
    ].sum()
    summary["예산(계획)_증감(B-C)"] = summary["예산(계획)_당기(B)"] - summary["예산(계획)_전기(C)"]
    summary["예산(계획)_증감율(%)"] = summary.apply(
        lambda row: (row["예산(계획)_증감(B-C)"] / abs(row["예산(계획)_전기(C)"])) * 100
        if row["예산(계획)_전기(C)"] != 0 else None, axis=1)

    pivot = df.pivot_table(
        index=["사업장_사업유형", "사업장_권역", "기관"],
        columns="매핑",
        values="예산(계획)_당기(B)",
        aggfunc="sum"
    ).fillna(0)

    order = ["인건비", "업무추진비", "운영비", "시설비", "사업비", "잡지출", "기타"]
    pivot = pivot.reindex(columns=order, fill_value=0).reset_index()
    result = pd.merge(summary, pivot, on=["사업장_사업유형", "사업장_권역", "기관"], how="left")
    return result[result[order].sum(axis=1) != 0]


def summarize_income_by_fund(df_income: pd.DataFrame) -> pd.DataFrame:
    income_mapping = {
        "보조금": "경상보조금", "비지정후원금": "후원금", "지정후원금": "후원금",
        "전년도 이월금(보조금)": "이월금", "전년도 이월금(법인)": "이월금", "전년도 이월금(후원금)": "이월금",
        "전년도 이월금(기타)": "이월금", "이자수입(본부)": "잡수입", "잡수입": "잡수입",
        "법인전입금(후원금)": "법인전입금(후원금)", "수증물품": "후원물품",
        "법인전입금(후원금) 기타": "법인전입금(후원금)기타", "전년도이월금(법인후원금_지역기타)": "이월금",
        "기타예금이자": "잡수입", "기타예금이자(후원금)": "잡수입",
        "전년도 이월금(법인후원금기타)": "이월금", "사업수입": "잡수입", "사용자부담금": "잡수입",
        "국고보조금": "경상보조금", "시도보조금": "경상보조금", "지정후원금(기타)": "법인전입금(지역)",
        "시군구보조금": "경상보조금", "전년도이월금(지정후원금기타)": "이월금", "기타보조금": "기타보조금",
        "매핑없음": "매핑없음", "없음": "매핑없음"
    }
    df = df_income.copy()
    df["기관"] = df.apply(lambda row: row["사업장_예산부서"] if row["사업장_권역"] == "본부" else row["기관"], axis=1)
    df["매핑"] = df["재원"].map(income_mapping)

    summary = df.groupby(["사업장_사업유형", "사업장_권역", "기관"], as_index=False)[
        ["예산(계획)_전기(C)", "예산(계획)_당기(B)"]
    ].sum()
    summary["예산(계획)_증감(B-C)"] = summary["예산(계획)_당기(B)"] - summary["예산(계획)_전기(C)"]
    summary["예산(계획)_증감율(%)"] = summary.apply(
        lambda row: (row["예산(계획)_증감(B-C)"] / abs(row["예산(계획)_전기(C)"])) * 100
        if row["예산(계획)_전기(C)"] != 0 else None, axis=1)

    pivot = df.pivot_table(
        index=["사업장_사업유형", "사업장_권역", "기관"],
        columns="매핑",
        values="예산(계획)_당기(B)",
        aggfunc="sum"
    ).fillna(0)

    order = ["경상보조금", "기타보조금", "후원금", "후원물품",
             "법인전입금(후원금)기타", "법인전입금(지역)", "법인전입금(후원금)",
             "잡수입", "이월금", "매핑없음"]
    pivot = pivot.reindex(columns=order, fill_value=0).reset_index()
    result = pd.merge(summary, pivot, on=["사업장_사업유형", "사업장_권역", "기관"], how="left")
    return result[result[order].sum(axis=1) != 0]


def apply_postprocessing_with_style(df: pd.DataFrame, sheet_name: str, writer: pd.ExcelWriter):
    df = df[df["기관"].notna() & (df["기관"].astype(str).str.strip() != "")]
    numeric_cols = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])]

    result_rows = []
    style_flags = []

    for name, group in df.groupby("사업장_사업유형", sort=True):
        group_sorted = group.sort_values(by=["사업장_권역", "기관"])
        result_rows.append(group_sorted)
        style_flags.extend([""] * len(group_sorted))

        subtotal = group[numeric_cols].sum(numeric_only=True)
        subtotal_row = {
            "사업장_사업유형": name,
            "사업장_권역": "소계",
            "기관": "소계"
        }
        subtotal_row.update(subtotal.to_dict())
        result_rows.append(pd.DataFrame([subtotal_row]))
        style_flags.append("소계")

    total = df[numeric_cols].sum(numeric_only=True)
    total_row = {
        "사업장_사업유형": "총계",
        "사업장_권역": "총계",
        "기관": "총계"
    }
    total_row.update(total.to_dict())
    result_rows.append(pd.DataFrame([total_row]))
    style_flags.append("총계")

    df_final = pd.concat(result_rows, ignore_index=True)
    df_final.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.book[sheet_name]
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for idx, flag in enumerate(style_flags, start=2):
        if flag in ["소계", "총계"]:
            for cell in ws[idx]:
                cell.fill = gray_fill


# -------- Colab UI 런처 함수 --------
def launch_ui():
    def handle_upload(btn):
        with output:
            clear_output()
            print("📤 업로드할 파일을 선택하세요.")
            uploaded = files.upload()
            if uploaded:
                clear_output()
                print("✅ 업로드 완료!")
                files_list = [f for f in os.listdir() if f.endswith(".xlsx") or f.endswith(".xls")]
                filename_dropdown.options = files_list
                run_button.disabled = False

    def handle_explore(btn):
        with output:
            clear_output()
            print("📁 Colab 내 엑셀 파일을 탐색 중...")
            files_list = [f for f in os.listdir() if f.endswith(".xlsx") or f.endswith(".xls")]
            if not files_list:
                print("❌ Colab 내 엑셀 파일이 없습니다.")
            else:
                print("✅ 파일 목록:", files_list)
                filename_dropdown.options = files_list
                filename_dropdown.value = files_list[0]
                run_button.disabled = False

    def run_processing(btn):
        with output:
            clear_output()
            try:
                file_name = filename_dropdown.value
                print(f"📄 선택된 파일: {file_name}")
                df_expense = pd.read_excel(file_name, sheet_name="시트1_지출재원항매핑")
                df_income = pd.read_excel(file_name, sheet_name="시트2_수입재원매핑")

                df1 = summarize_expense_by_account(df_expense)
                df2 = summarize_income_by_fund(df_income)

                output_file = f"요약_{file_name}"
                with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                    apply_postprocessing_with_style(df1, "지출요약", writer)
                    apply_postprocessing_with_style(df2, "수입요약", writer)

                print("✅ 요약 완료! 파일 다운로드 시작...")
                files.download(output_file)
            except Exception as e:
                print("❌ 오류 발생:", e)

    upload_button.on_click(handle_upload)
    explore_button.on_click(handle_explore)
    run_button.on_click(run_processing)

    display(widgets.VBox([
        widgets.HBox([upload_button, explore_button]),
        filename_dropdown,
        run_button,
        output
    ]))
